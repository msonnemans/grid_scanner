from PIL import Image
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import matplotlib.colors as mcolors
import cv2
import operator
from openpyxl import Workbook
from os import listdir
from os.path import isfile, join

image = []
baw_array = []
row = 0
col = 0
threshold = 80
num_boxes = 60
grid_rows = 10
grid_columns = 4
total_columns = 10
total_rows = 6
percentage_black_threshold = 0.3
line_width_factor = 0.05
colors = ['g', 'y', 'b', 'r']

def find_boxes():
    boxes = []
    visited = set()
    for i in range(row):
        for j in range(col):
            if (i, j) in visited or image.item(i, j) >= threshold:
                continue
            coord = _flood_fill(i, j, visited)
            if len(boxes) < num_boxes:
                boxes.append(coord)
                boxes.sort(key=box_size, reverse=True)
            else:
                boxes.append(coord)
                boxes.sort(key=box_size, reverse=True)
                boxes.pop()
    return boxes

def box_size(coord):
    a1, b1, a2, b2 = coord
    return (a2 - a1) * (b2 - b1)

def _flood_fill(i, j, visited):
    r1 = r2 = i
    c1 = c2 = j
    stack = [(i, j)]
    visited.add((i, j))
    while stack:
        r, c = stack.pop()
        for nr, nc in [(r+1, c), (r-1, c), (r, c+1), (r, c-1)]:
            if nr < 0 or nc < 0 or nr >= row or nc >= col:
                continue
            if (nr, nc) in visited or image.item(nr, nc) >= threshold:
                continue
            r1 = min(nr, r1)
            r2 = max(nr, r2)
            c1 = min(nc, c1)
            c2 = max(nc, c2)
            stack.append((nr, nc))
            visited.add((nr, nc))
    return (c1, r1, c2, r2)

def get_y_pos(box):
    a1, b1, a2, b2 = box
    return b1

def get_x_pos(box):
    a1, b1, a2, b2 = box
    return a1


def get_sub_box(box, column, row):
    x1, y1, x2, y2 = box
    width = x2 - x1
    height = y2 - y1

    line_width = width * line_width_factor

    x1 = x1 + line_width
    y1 = y1 + line_width
    x2 = x2 - line_width
    y2 = y2 - line_width
    width = width - (line_width * 2)
    height = height - (line_width * 2)

    box_width = width / grid_columns
    box_height = height / grid_rows

    mini_box_x1 = x1 + (box_width * column) + line_width
    mini_box_y1 = y1 + (box_height * row) + line_width
    mini_box_x2 = x1 + (box_width * column) + box_width - line_width
    mini_box_y2 = y1 + (box_height * row) + box_height - line_width

    return (mini_box_x1, mini_box_y1, mini_box_x2, mini_box_y2)

def get_pixels_at_pos(box, column, row):
    x1, y1, x2, y2 = get_sub_box(box, column, row)

    pixels = []
    for x in range(int(x1), int(x2)):
        for y in range(int(y1), int(y2)):
            pixels.append(baw_array.item(y, x))
    return pixels

def decide_color(pixel):
    if pixel > 200:
        return 0
    else:
        return 1

def get_black_percentage(pixels):
    if not pixels:
        return 0
    mapped = list(map(decide_color, pixels))
    return sum(mapped) / len(mapped)

def chunks(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

def max_key_value(dic):
    return max(dic, key=lambda key: dic[key])

def decide_grid(grid):
    sub_grid = {}
    for y in range(grid_rows):
        row = {}
        for x in range(grid_columns):
            perc = get_black_percentage(get_pixels_at_pos(grid, x, y))
            if perc > percentage_black_threshold:
                row[x] = perc
        if len(row) > 0:
            sub_grid[y] = max_key_value(row)
    return sub_grid

def draw_grid(_ax, grid):
    decided_grid = decide_grid(grid)
    for y, x in decided_grid.items():
        x1, y1, x2, y2 = get_sub_box(grid, x, y)
        rect = patches.Rectangle((x1,y1),x2-x1,y2-y1,fill=True)
        ax.add_patch(rect)

def calculate_checked_boxes():
    grids = find_boxes()
    grids.sort(key=get_y_pos)
    full_grid = []
    for row in chunks(grids, total_columns):
        row.sort(key=get_x_pos)
        full_grid.append(row)
    return full_grid

def get_color_from_row_index(row):
    return colors[row]

def get_leading_zero(num):
    if num < 10 :
        return str("0") + str(num)
    else:
        return str(num)

all_files = [f for f in listdir('.') if isfile(f)]
all_images = [f for f in all_files if f.endswith(".jpg") or f.endswith(".jpeg") or f.endswith(".png")]
all_hours = [f for f in all_images if f.split('.')[0].isnumeric()]

book = Workbook()
sh = book.active
for i, filename in enumerate(all_hours):
    hour = int(filename.split('.')[0])
    for n in range(60):
        sh.cell((i * 60) + n+1, 1, get_leading_zero(hour) + ":" + get_leading_zero(n))
    
    img = Image.open(filename)
    img = img.convert('L')
    image = np.array(img)

    img = cv2.imread(filename)
    img = cv2.GaussianBlur(img,(5,5),cv2.BORDER_DEFAULT)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    (thresh, baw) = cv2.threshold(img, 180, 255, cv2.THRESH_BINARY)

    baw_array = np.array(baw)
    row, col = image.shape

    full_grid = calculate_checked_boxes()
    for n in range(total_columns):   
        for m in range(total_rows):
            for y, x in decide_grid(full_grid[m][n]).items():
                sh.cell((i * 60) + (m*10)+y+1, n+2, get_color_from_row_index(x))

book.save("output.xlsx")

# img = Image.open('08.jpg')
# img = img.convert('L')
# image = np.array(img)

# img = cv2.imread('08.jpg')
# img = cv2.GaussianBlur(img,(5,5),cv2.BORDER_DEFAULT)
# img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
# (thresh, baw) = cv2.threshold(img, 180, 255, cv2.THRESH_BINARY)

# baw_array = np.array(baw)
# row, col, = image.shape

# fig,ax = plt.subplots(1)
# ax.imshow(image, cmap='gray')

# full_grid = calculate_checked_boxes()
# for row in full_grid:
#     for box in row:
#         x1, y1, x2, y2 = box
#         rect = patches.Rectangle((x1,y1),x2-x1,y2-y1,fill=False,lw=1,color=mcolors.CSS4_COLORS['lime'])
#         ax.add_patch(rect)
# for n in range(total_columns):
#     for m in range(total_rows):
#         draw_grid(ax, full_grid[m][n])

# plt.show()
