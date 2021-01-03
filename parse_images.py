from openpyxl import Workbook
from PIL import Image
import numpy as np
import cv2
from box_utils import find_boxes

def get_leading_zero(num):
    if num < 10 :
        return str("0") + str(num)
    else:
        return str(num)

def get_y_pos(box):
    a1, b1, a2, b2 = box
    return b1

def get_x_pos(box):
    a1, b1, a2, b2 = box
    return a1

def get_sub_box(box, column, row, line_width_factor, grid_columns, grid_rows):
    x1, y1, x2, y2 = box
    width = x2 - x1
    height = y2 - y1

    line_width = width * line_width_factor

    x1 = x1 + line_width
    y1 = y1 + line_width
    x2 = x2 - line_width
    y2 = y2 - line_width
    width = width - (line_width * 2)
    height = height - (line_width * 2)

    box_width = width / grid_columns
    box_height = height / grid_rows

    mini_box_x1 = x1 + (box_width * column) + line_width
    mini_box_y1 = y1 + (box_height * row) + line_width
    mini_box_x2 = x1 + (box_width * column) + box_width - line_width
    mini_box_y2 = y1 + (box_height * row) + box_height - line_width

    return (mini_box_x1, mini_box_y1, mini_box_x2, mini_box_y2)

def get_pixels_at_pos(box, column, row, line_width_factor, grid_columns, grid_rows, baw_array):
    x1, y1, x2, y2 = get_sub_box(box, column, row, line_width_factor, grid_columns, grid_rows)

    pixels = []
    for x in range(int(x1), int(x2)):
        for y in range(int(y1), int(y2)):
            pixels.append(baw_array.item(y, x))
    return pixels

def chunks(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

def decide_color(pixel):
    if pixel > 200:
        return 0
    else:
        return 1

def max_key_value(dic):
    return max(dic, key=lambda key: dic[key])

def get_black_percentage(pixels):
    if not pixels:
        return 0
    mapped = list(map(decide_color, pixels))
    return sum(mapped) / len(mapped)

def calculate_checked_boxes(row, col, threshold, num_boxes, image, total_columns):
    grids = find_boxes(row, col, threshold, num_boxes, image)
    grids.sort(key=get_y_pos)
    full_grid = []
    for row in chunks(grids, total_columns):
        row.sort(key=get_x_pos)
        full_grid.append(row)
    return full_grid

def decide_grid(grid, grid_rows, grid_columns, line_width_factor, baw_array, percentage_black_threshold):
    sub_grid = {}
    for y in range(grid_rows):
        row = {}
        for x in range(grid_columns):
            perc = get_black_percentage(get_pixels_at_pos(grid, x, y, line_width_factor, grid_columns, grid_rows, baw_array))
            if perc > percentage_black_threshold:
                row[x] = perc
        if len(row) > 0:
            sub_grid[y] = max_key_value(row)
    return sub_grid

### fileList: list of full paths to images
### hour: starting hour
### colors: ordered list of cell color value
### total columns: the total amount of large box columns
### total_rows: the total amount of large box rows
### grid_rows: the total amount of rows per single large box
### grid_columns: the total amount of columns per single large box
### threshold: edge detection threshold (0-255). default: 80
### line_width_factor: the factor width which line widths are multiplied to find the optimal boxes
### percentage_black_threshold: percentage of black pixels to decide a small cell to be black
def parse(fileList, hour, colors, total_columns, total_rows, grid_rows, grid_columns, threshold = 80, line_width_factor = 0.05, percentage_black_threshold = 0.3):
    book = Workbook()
    sh = book.active

    for i, filename in enumerate(fileList):
        for n in range(60):
            sh.cell((i * 60) + n+1, 1, get_leading_zero(hour) + ":" + get_leading_zero(n))
        
        img = Image.open(filename)
        img = img.convert('L')
        image = np.array(img)
        img.close()

        img = cv2.imread(filename)
        img = cv2.GaussianBlur(img,(5,5),cv2.BORDER_DEFAULT)
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        (thresh, blackAndWhite) = cv2.threshold(img, 180, 255, cv2.THRESH_BINARY)

        blackAndWhiteArray = np.array(blackAndWhite)
        row, col = image.shape

        full_grid = calculate_checked_boxes(row, col, threshold, total_columns * total_rows, image, total_columns)
        for n in range(total_columns):   
            for m in range(total_rows):
                for y, x in decide_grid(full_grid[m][n], grid_rows, grid_columns, line_width_factor, blackAndWhiteArray, percentage_black_threshold).items():
                    sh.cell((i * 60) + (m*10)+y+1, n+2, colors[x])

    return book